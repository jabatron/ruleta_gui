from openpyxl.workbook import workbook
import whatsapp as w
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from time import time
from time import sleep
import gspread
import re
import argparse
import os.path
import sys

def enviar_xls(messenger, args):
 
    # Abrir excel
    wb = load_workbook(file)
    sheet = wb.active
 
    # Procesar argumentos
    enviar_mensaje = args.m
    enviar_fichero = args.s
    if args.f:
        file = args.f
    else:
        file = 'moviles.xlsx'

    for i in range (2, sheet.max_row + 1):
        cell_numero  = 'A' + str(i)
        cell_mensaje = 'B' + str(i)
        cell_fichero = 'C' + str(i)
        cell_qsoy    = 'D' + str(i)
        cell_status  = 'G' + str(i)
        numero  = '34' + str(sheet[cell_numero].value)[:]
        pattern = re.compile("^34[67]\d{8}")
        mensaje = sheet[cell_mensaje].value
        fichero = sheet[cell_fichero].value
        qsoy    = sheet[cell_qsoy].value
        
        
        if qsoy != args.q and args.q != None :
            print (f"no soy yo {qsoy}")
            continue
        
               
        if pattern.fullmatch(numero):
            
            if enviar_mensaje:

                if not args.d:
                    messenger.find_user(numero)
                    messenger.send_message(mensaje)

                print (f'Enviando mensaje: -{mensaje}- al numero {numero}')

            if enviar_fichero:
                # comprobar si ha fichero para enviar y existe
                # sacar tipo de fichero
                # enviar segun tipo de fichero
                if fichero and os.path.isfile(fichero): 
                    name, ext = os.path.splitext(fichero)
                    
                    if not args.d:
                        messenger.find_user(numero)
                    
                    if ext.lower() in [".tiff", ".pjp", ".jfif", ".gif", ".svg", ".bmp", ".png", ".jpeg", ".svgz", ".jpg", ".webp", ".ico", ".xbm", ".dib", ".tif", ".pjpeg", ".avif", ".m4v", ".mp4", ".3gpp", ".mov"]:
                        if not args.d:
                            messenger.send_picture(fichero)
                        print (f'Enviando imagen/video {fichero} al número: {numero}')
                    else:
                        if not args.d:
                            messenger.send_file(fichero)
                        print (f'Enviando fichero {fichero} al número: {numero}')
                else:
                    print (f'No hay fichero para enviar al numero {numero}')

            sheet[cell_status].font = Font(color="00F000")
            sheet[cell_status] = 'Enviado'
            if not args.d:
                wb.save(file)

        else:
            print (f'el numero {numero} es incorrecto')
            sheet[cell_status].font = Font(color="FF0000")
            sheet[cell_numero].font = Font(color="FF0000")
            sheet[cell_status] = 'ERROR'
            if not args.d:
                wb.save(file)

def enviar_google(messenger, args):

    # Abrir gooogle sheet
    gc = gspread.service_account(filename="credentials/ruleta-gui-credentials.json")
    key = open("credentials/key-mobiles.txt").read()

    try:
        worksheet = gc.open_by_key(key).sheet1
    except:
        print ('Error con el fichero de Google Sheet')
        sys.exit()
    
    
    data = worksheet.get_all_values()
    #Leer cabeceras que no necesitamos para nada 
    cabeceras = data.pop(0)

    # Procesar argumentos
    enviar_mensaje = args.m
    enviar_fichero = args.s

    for i, d in enumerate(data, 2):
        numero, mensaje, fichero, qsoy, *_ = d
        numero = '34' + numero
        pattern = re.compile("^34[67]\d{8}")
   
        if qsoy != args.q and args.q != None :
            print (f"no soy yo {qsoy}")
            continue
                   
        if pattern.fullmatch(numero):    
            if enviar_mensaje:
                if not args.d:
                    messenger.find_user(numero)
                    messenger.send_message(mensaje)
                    worksheet.update('G' + str(i), 'Mensaje Enviado')
                    worksheet.format('G' + str(i), {"textFormat": {"foregroundColor": {"red": 0.0, "green": 75.0, "blue": 0.0}}})

                print (f'Enviando mensaje: -{mensaje}- al numero {numero}')

            if enviar_fichero:
                # comprobar si ha fichero para enviar y existe
                # sacar tipo de fichero
                # enviar segun tipo de fichero
                if fichero and os.path.isfile(fichero): 
                    name, ext = os.path.splitext(fichero)
                    
                    if not args.d:
                        messenger.find_user(numero)
                    
                    if ext.lower() in [".tiff", ".pjp", ".jfif", ".gif", ".svg", ".bmp", ".png", ".jpeg", ".svgz", ".jpg", ".webp", ".ico", ".xbm", ".dib", ".tif", ".pjpeg", ".avif", ".m4v", ".mp4", ".3gpp", ".mov"]:
                        if not args.d:
                            messenger.send_picture(fichero)
                            worksheet.update('H' + str(i), 'Imagen/Video Enviado')
                            worksheet.format('H' + str(i), {"textFormat": {"foregroundColor": {"red": 0.0, "green": 75.0, "blue": 0.0}}})
                        print (f'Enviando imagen/video {fichero} al número: {numero}')
                    else:
                        if not args.d:
                            messenger.send_file(fichero)
                            worksheet.update('H' + str(i), 'Fichero Enviado')
                            worksheet.format('H' + str(i), {"textFormat": {"foregroundColor": {"red": 0.0, "green": 75.0, "blue": 0.0}}})
                        print (f'Enviando fichero {fichero} al número: {numero}')
                else:
                    print (f'No hay fichero para enviar al numero {numero}')

        else:
            if not args.d and args.g:
                worksheet.update('H' + str(i), 'No Enviado')
                worksheet.format('H' + str(i), {"textFormat": {"foregroundColor": {"red": 50.0, "green": 0.0, "blue": 0.0}}})
            if not args.d and args.s:
                worksheet.update('G' + str(i), 'No Enviado')
                worksheet.format('G' + str(i), {"textFormat": {"foregroundColor": {"red": 50.0, "green": 0.0, "blue": 0.0}}})
            print (f'el numero {numero} es incorrecto')

def check_file(file):
    try:
        libro = load_workbook(file)
        return file
    except:
        msg = f'{file} No existe o no es Excel.'
        raise argparse.ArgumentTypeError(msg)


def main():
    
    parser = argparse.ArgumentParser (description="Envio de mensajes por WhastApp/web", epilog="@jabaselga")
    group = parser.add_mutually_exclusive_group()
    group.add_argument("-f",  metavar="moviles.xlsx", help="Fichero de Excel con los datos.", type=check_file)
    group.add_argument("-g",  help="Fichero de Google con los datos.", action="store_true")
    parser.add_argument("-m", help="Enviar mensaje", action="store_true")
    parser.add_argument("-s", help="Enviar imagen/video/fichero", action="store_true")
    parser.add_argument("-q", metavar="string", help="Quien soy")
    parser.add_argument("-d", help="Simulación, no enviar mensajes ni archivos", action="store_true")
    parser.add_argument("-v", help="Extra info", action="store_true")
    
    args = parser.parse_args()   
    
    if not (args.m or args.s):
        parser.error('Ninguna tarea por hacer sleccione "-m" y/o "-s"')

    messenger = w.WhatsApp()

    t1 = time()

    if args.g:
        enviar_google(messenger, args)
    else:
        enviar_xls(messenger, args)


    # Añado un delay para asegurarme que ee han enviado todos los mensajes    
    sleep(3)
    t2 = time()

    print (f'El proceso ha durado {t2-t1} segundos.')

if __name__=='__main__':
    main ()



